"""
Excel-parser för svenska mängdförteckningar (.xlsx, .xlsm).

Hanterar både rena MF (Westcon-stil med kod på varje rad) och
TB+MF-kombinationer (Trafikverket/Sweco-stil där AMA-koden står
i en sektionsrad ovanför textraderna).

Återanvänder OfferLine/OfferDocument från parser.py.
"""

from __future__ import annotations

import io
import re
import unicodedata
from typing import Any

from openpyxl import load_workbook

from app.parser import (
    OfferDocument,
    OfferLine,
    is_ama_code,
    parse_swedish_number,
)


# Whitespace-normalisering för header/label-matching.
# Excel-headers innehåller ofta U+2004 (three-per-em space) eller andra
# unicode-whitespace som bryter exakt strängmatchning.
def _norm_label(value: Any) -> str:
    if value is None:
        return ""
    s = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", " ", s).strip().upper()


def _to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return parse_swedish_number(str(value))


def parse_excel_bytes(data: bytes) -> OfferDocument:
    """Parse Excel-MF till OfferDocument. Försöker varje blad."""
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)

    last_error: Exception | None = None
    for ws in wb.worksheets:
        try:
            rows = _read_sheet_rows(ws, max_rows=10_000)
            if not rows:
                continue
            doc = _parse_excel_rows(rows)
            # Acceptera bara om vi faktiskt hittade några rader
            if doc.lines:
                return doc
        except Exception as e:
            last_error = e
            continue

    raise ValueError(
        "Hittade ingen mängdförteckning-struktur i Excel-filen"
        + (f" (senaste fel: {last_error})" if last_error else "")
    )


def _read_sheet_rows(ws, max_rows: int = 10_000) -> list[list[Any]]:
    """Läs rader som lista av lista. Behåller native typer."""
    out: list[list[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        cells = list(row)
        while cells and cells[-1] is None:
            cells.pop()
        out.append(cells)
    return out


# Header-tokens vi letar efter, normaliserade
_HEADER_TOKENS = {
    "kod":        {"KOD", "AMA-KOD", "AMA"},
    "text":       {"TEXT", "BENÄMNING", "BESKRIVNING"},
    "unit":       {"ENHET"},
    "qty":        {"MÄNGD", "ANTAL"},
    "unit_price": {"À-PRIS", "Á-PRIS", "A-PRIS", "ENHETSPRIS", "À PRIS", "À-PRIS (KR)"},
    "total":      {"BELOPP", "TOTALT", "SUMMA", "TOTALSUMMA"},
}


def _detect_columns(rows: list[list[Any]]) -> dict[str, int]:
    """Letar efter header-rad i de första 30 raderna."""
    for row_idx, row in enumerate(rows[:30]):
        norm_row = [_norm_label(c) for c in row]
        if any(c in _HEADER_TOKENS["kod"] for c in norm_row) and \
           any(c in _HEADER_TOKENS["text"] for c in norm_row):
            mapping: dict[str, int] = {}
            for key, tokens in _HEADER_TOKENS.items():
                for col_idx, val in enumerate(norm_row):
                    if val in tokens:
                        mapping[key] = col_idx
                        break
            mapping["_header_row"] = row_idx
            return mapping
    raise ValueError("Hittade ingen kolumnheader (KOD/TEXT/…) i Excel-arket")


_METADATA_LABELS = {
    "PROJEKT": "project_name",
    "OBJEKT": "project_name",
    "DOKUMENTNUMMER": "document_number",
    "DOKUMENTNR": "document_number",
    "HANDLINGSNR": "document_number",
    "DATUM": "date",
    "HANDLÄGGARE": "handlaggare",
    "UPPDRAGSNUMMER": "uppdragsnummer",
    "UPPDRAGSNR": "uppdragsnummer",
    "TOTALT BELOPP": "total_amount",
    "TOTALSUMMA": "total_amount",
}


def _extract_metadata(rows: list[list[Any]]) -> dict:
    """Plocka projekt/dokument-metadata. Letar BARA på raden NEDANFÖR
    label-cellen — undviker att felaktigt plocka 'DOKUMENTNUMMER' som
    värde för 'PROJEKT' när labels står på samma rad."""
    meta: dict = {
        "project_name": None,
        "document_number": None,
        "date": None,
        "handlaggare": None,
        "uppdragsnummer": None,
        "total_amount": None,
        "status": None,
    }

    for row_idx, row in enumerate(rows[:15]):
        for col_idx, cell in enumerate(row):
            label = _norm_label(cell)
            target = _METADATA_LABELS.get(label)
            if not target:
                continue
            # Värdet är ALLTID på raden under, samma kolumn
            if row_idx + 1 < len(rows):
                next_row = rows[row_idx + 1]
                if col_idx < len(next_row):
                    raw = next_row[col_idx]
                    value = _to_str(raw)
                    if value and value not in {"-", ":"}:
                        if target == "total_amount":
                            if isinstance(raw, (int, float)) and not isinstance(raw, bool):
                                meta[target] = float(raw)
                            else:
                                meta[target] = parse_swedish_number(value)
                        elif meta[target] is None:
                            meta[target] = value

    return meta


def _parse_excel_rows(rows: list[list[Any]]) -> OfferDocument:
    meta = _extract_metadata(rows)
    cols = _detect_columns(rows)
    header_row = cols["_header_row"]

    doc = OfferDocument(
        document_number=meta["document_number"],
        project_name=meta["project_name"],
        date=meta["date"],
        handlaggare=meta["handlaggare"],
        uppdragsnummer=meta["uppdragsnummer"],
        total_amount=meta["total_amount"],
        status=meta["status"],
    )

    current_ama_code: str | None = None
    current_section_title: str | None = None
    sum_amount = 0.0

    max_col = max(cols.get(k, 0) for k in ["kod", "text", "unit", "qty", "unit_price", "total"])

    for idx, row in enumerate(rows[header_row + 1:], start=header_row + 2):
        if len(row) <= max_col:
            row = row + [None] * (max_col - len(row) + 1)

        kod = _to_str(row[cols["kod"]]) if "kod" in cols else ""
        text = _to_str(row[cols["text"]]) if "text" in cols else ""
        unit = _to_str(row[cols["unit"]]) if "unit" in cols else ""
        qty_raw = row[cols["qty"]] if "qty" in cols else None
        price_raw = row[cols["unit_price"]] if "unit_price" in cols else None
        total_raw = row[cols["total"]] if "total" in cols else None

        qty = _to_number(qty_raw)
        unit_price = _to_number(price_raw)
        total = _to_number(total_raw)

        # Fält som indikerar att raden är "MF-aktig"
        has_unit = bool(unit) and unit != "-"
        has_qty = qty is not None and qty != 0
        has_price = unit_price is not None and unit_price != 0
        has_total = total is not None and total != 0

        # 1. Helt tom rad → skip
        if not any([kod, text, has_unit, has_qty, has_price, has_total]):
            continue

        # 2. Sektion-rubrik: AMA-kod i kod-kolumn, ingen mängd/pris
        if is_ama_code(kod) and not (has_qty or has_price or has_total):
            current_ama_code = kod
            current_section_title = text or current_section_title
            continue

        # 3. TB-text: bara text, ingen enhet/mängd/pris/kod → skip
        if not is_ama_code(kod) and not has_unit and not has_qty and not has_price and not has_total:
            continue

        # 4. Riktig MF-rad: har enhet+mängd ELLER egen AMA-kod
        if not text:
            continue

        unit_str = unit if has_unit else None
        is_lump = (
            (not unit_str)
            and qty is None
            and unit_price is None
            and total is not None
        )

        # Använd radens egen kod om den är giltig AMA-kod, annars sektion-koden
        line_code = kod if is_ama_code(kod) else current_ama_code

        line = OfferLine(
            line_number=idx,
            ama_code=line_code,
            ama_section_title=current_section_title,
            description=text,
            unit=unit_str,
            quantity=qty,
            unit_price=unit_price,
            total_amount=total,
            is_lump_sum=is_lump,
            raw_row=[_to_str(c) for c in row],
        )
        doc.lines.append(line)
        if total:
            sum_amount += total

    # Om metadata inte hade totalsumma — använd summan av rader (kan vara 0
    # i en tom MF där anbudsgivaren ska fylla i priserna)
    if doc.total_amount is None and sum_amount > 0:
        doc.total_amount = round(sum_amount, 2)

    if not doc.lines:
        raise ValueError("Hittade inga rader under header — bladet kan vara tomt eller ha annan struktur")

    return doc
