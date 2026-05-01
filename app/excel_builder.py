"""
Excel-mall (Lodet output) — strukturerad mängdförteckning med AI-föreslagna
priser, konfidensindikatorer och länkar till historiska matchningar.
"""

from __future__ import annotations

import io
import random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LODBLA = "1C2C3A"
MURSAND = "F5F1EA"
TEGEL = "B8572E"
OCKRA = "D4A017"
SALVIA = "6B8E5B"
JARN = "2B2B2B"
KRITA = "FFFFFF"
LJUSGRA = "EDE9E1"


def thin_border(color: str = "D4CFC4") -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def style_header_cell(cell, bg_color: str = LODBLA, fg_color: str = KRITA):
    cell.font = Font(name="Inter", size=10, bold=True, color=fg_color)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border()


def style_data_cell(cell, bold: bool = False, mono: bool = False, color: str = JARN):
    font_name = "JetBrains Mono" if mono else "Inter"
    cell.font = Font(name=font_name, size=10, bold=bold, color=color)
    cell.border = thin_border()
    if mono:
        cell.alignment = Alignment(horizontal="right", vertical="center")
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def simulate_ai_suggestion(original_price: float | None) -> dict:
    if original_price is None:
        return {
            "suggested_price": None,
            "confidence": "låg",
            "match_count": 0,
            "historical_range": None,
            "rationale": "Ingen historisk match — manuell prissättning krävs",
        }

    index_factor = 1.0 + random.uniform(0.02, 0.06)
    suggested = round(original_price * index_factor, 0)

    spread = random.uniform(0.05, 0.20)
    p25 = round(suggested * (1 - spread / 2), 0)
    p75 = round(suggested * (1 + spread / 2), 0)
    match_count = random.choice([2, 3, 4, 5, 6, 8])

    if match_count >= 5 and spread < 0.10:
        confidence = "hög"
    elif match_count >= 3:
        confidence = "medel"
    else:
        confidence = "låg"

    return {
        "suggested_price": suggested,
        "confidence": confidence,
        "match_count": match_count,
        "historical_range": (p25, p75),
        "rationale": f"{match_count} historiska träffar, indexjusterat E84",
    }


def build_anbud_sheet(wb: Workbook, doc_data: dict, generated_at: str) -> None:
    sheet = wb.active
    sheet.title = "Anbud"
    sheet.sheet_view.showGridLines = False

    meta = doc_data["metadata"]
    lines = doc_data["lines"]

    sheet["A1"] = "LODET"
    sheet["A1"].font = Font(name="Inter", size=18, bold=True, color=LODBLA)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 30

    sheet["A2"] = "Anbudsförslag — utkast för granskning"
    sheet["A2"].font = Font(name="Inter", size=10, italic=True, color="6B6B6B")

    sheet["A4"] = "Projekt"
    sheet["B4"] = meta.get("project_name") or "—"
    sheet["A5"] = "Dokumentnr"
    sheet["B5"] = meta.get("document_number") or "—"
    sheet["A6"] = "Status"
    sheet["B6"] = meta.get("status") or "—"
    sheet["A7"] = "Datum"
    sheet["B7"] = meta.get("date") or "—"
    sheet["A8"] = "Genererat av Lodet"
    sheet["B8"] = generated_at

    for r in range(4, 9):
        sheet.cell(row=r, column=1).font = Font(name="Inter", size=10, bold=True, color=LODBLA)
        sheet.cell(row=r, column=2).font = Font(name="Inter", size=10, color=JARN)

    headers = [
        ("AMA-kod", 12),
        ("Beskrivning", 45),
        ("Enhet", 7),
        ("Antal", 9),
        ("À-pris (förslag)", 14),
        ("Belopp (förslag)", 15),
        ("Konf.", 8),
        ("Träffar", 9),
        ("À-pris (justerat)", 15),
        ("Belopp (justerat)", 15),
        ("Kommentar", 30),
    ]

    header_row = 11
    for col_idx, (h, width) in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_idx, value=h)
        style_header_cell(cell)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    sheet.row_dimensions[header_row].height = 28

    current_section: str | None = None
    row = header_row + 1

    for line in lines:
        section_letter = (line.get("ama_code") or "")[:1]
        if section_letter != current_section and section_letter:
            section_label = {
                "B": "B — Förarbeten, hjälparbeten, saneringsarbeten",
                "C": "C — Mark- och anläggningsarbeten",
                "D": "D — Markförstärkningar och bärande konstruktioner",
                "E": "E — Konstruktionsarbeten",
                "S": "S — Apparater, ledningar m.m. i el- och telesystem",
                "Y": "Y — Märkning, kontroll, dokumentation, mediaförsörjning",
            }.get(section_letter, f"{section_letter} — Övrigt")

            section_cell = sheet.cell(row=row, column=1, value=section_label)
            section_cell.font = Font(name="Inter", size=10, bold=True, color=KRITA)
            section_cell.fill = PatternFill("solid", fgColor=TEGEL)
            section_cell.alignment = Alignment(horizontal="left", vertical="center")
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
            sheet.row_dimensions[row].height = 22
            row += 1
            current_section = section_letter

        ai = simulate_ai_suggestion(line.get("unit_price"))

        sheet.cell(row=row, column=1, value=line.get("ama_code") or "—")
        sheet.cell(row=row, column=2, value=line.get("description") or "")
        sheet.cell(row=row, column=3, value=line.get("unit") or "—")
        sheet.cell(row=row, column=4, value=line.get("quantity"))
        sheet.cell(row=row, column=5, value=ai["suggested_price"])

        if line.get("quantity") is not None and ai["suggested_price"] is not None:
            sheet.cell(row=row, column=6, value=f"=D{row}*E{row}")
        else:
            sheet.cell(row=row, column=6, value=line.get("total_amount"))

        sheet.cell(row=row, column=7, value=ai["confidence"])
        sheet.cell(row=row, column=8, value=ai["match_count"] or "")
        sheet.cell(row=row, column=9, value=None)
        sheet.cell(row=row, column=10, value=f"=IF(I{row}<>\"\",I{row}*D{row},F{row})")
        sheet.cell(row=row, column=11, value=ai["rationale"])

        for col_idx in range(1, 12):
            cell = sheet.cell(row=row, column=col_idx)
            mono = col_idx in {1, 4, 5, 6, 8, 9, 10}
            style_data_cell(cell, mono=mono)

        sheet.cell(row=row, column=5).fill = PatternFill("solid", fgColor="FBF1D8")
        sheet.cell(row=row, column=6).fill = PatternFill("solid", fgColor="FBF1D8")

        conf_cell = sheet.cell(row=row, column=7)
        if ai["confidence"] == "hög":
            conf_cell.fill = PatternFill("solid", fgColor="DCE6D4")
            conf_cell.font = Font(name="Inter", size=10, bold=True, color=SALVIA)
        elif ai["confidence"] == "medel":
            conf_cell.fill = PatternFill("solid", fgColor="FBF1D8")
            conf_cell.font = Font(name="Inter", size=10, bold=True, color="8A6A0F")
        else:
            conf_cell.fill = PatternFill("solid", fgColor="F0DDD3")
            conf_cell.font = Font(name="Inter", size=10, bold=True, color=TEGEL)
        conf_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx in [4]:
            sheet.cell(row=row, column=col_idx).number_format = '#,##0;(#,##0);-'
        for col_idx in [5, 6, 9, 10]:
            sheet.cell(row=row, column=col_idx).number_format = '#,##0 "kr";(#,##0);-'

        row += 1

    total_row = row + 1
    sheet.cell(row=total_row, column=2, value="ANBUDSSUMMA")
    sheet.cell(row=total_row, column=2).font = Font(name="Inter", size=11, bold=True, color=LODBLA)
    sheet.cell(row=total_row, column=6, value=f"=SUM(F{header_row + 1}:F{row - 1})")
    sheet.cell(row=total_row, column=10, value=f"=SUM(J{header_row + 1}:J{row - 1})")

    for col_idx in [6, 10]:
        c = sheet.cell(row=total_row, column=col_idx)
        c.font = Font(name="JetBrains Mono", size=11, bold=True, color=LODBLA)
        c.number_format = '#,##0 "kr";(#,##0);-'
        c.fill = PatternFill("solid", fgColor=OCKRA)
        c.alignment = Alignment(horizontal="right")
        c.border = Border(top=Side(style="medium", color=LODBLA),
                          bottom=Side(style="medium", color=LODBLA))

    sheet.row_dimensions[total_row].height = 24
    sheet.freeze_panes = f"A{header_row + 1}"


def build_summary_sheet(wb: Workbook, doc_data: dict) -> None:
    sheet = wb.create_sheet("Sammanfattning")
    sheet.sheet_view.showGridLines = False

    sheet["A1"] = "SAMMANFATTNING"
    sheet["A1"].font = Font(name="Inter", size=18, bold=True, color=LODBLA)
    sheet.row_dimensions[1].height = 30

    sheet["A3"] = "Matchningskvalitet"
    sheet["A3"].font = Font(name="Inter", size=12, bold=True, color=LODBLA)

    stats = [
        ("Totalt antal rader", len(doc_data["lines"])),
        ("Rader med AI-förslag", len([l for l in doc_data["lines"] if l.get("unit_price") is not None])),
        ("Rader utan match (manuell)", len([l for l in doc_data["lines"] if l.get("unit_price") is None])),
        ("Klumpsummerader", len([l for l in doc_data["lines"] if l.get("is_lump_sum")])),
    ]

    for idx, (label, value) in enumerate(stats, start=4):
        sheet.cell(row=idx, column=1, value=label).font = Font(name="Inter", size=10, color=JARN)
        sheet.cell(row=idx, column=2, value=value).font = Font(
            name="JetBrains Mono", size=10, bold=True, color=LODBLA
        )
        sheet.cell(row=idx, column=2).alignment = Alignment(horizontal="right")

    sheet["A10"] = "Konfidensfördelning"
    sheet["A10"].font = Font(name="Inter", size=12, bold=True, color=LODBLA)

    sheet["A11"] = "Hög konfidens"
    sheet["B11"] = '= COUNTIF(Anbud!G:G, "hög")'
    sheet["A11"].font = Font(name="Inter", size=10, color=SALVIA, bold=True)

    sheet["A12"] = "Medel konfidens"
    sheet["B12"] = '= COUNTIF(Anbud!G:G, "medel")'
    sheet["A12"].font = Font(name="Inter", size=10, color="8A6A0F", bold=True)

    sheet["A13"] = "Låg konfidens"
    sheet["B13"] = '= COUNTIF(Anbud!G:G, "låg")'
    sheet["A13"].font = Font(name="Inter", size=10, color=TEGEL, bold=True)

    for r in [11, 12, 13]:
        sheet.cell(row=r, column=2).font = Font(
            name="JetBrains Mono", size=10, bold=True, color=LODBLA
        )
        sheet.cell(row=r, column=2).alignment = Alignment(horizontal="right")

    sheet["A16"] = "Anbudssumma"
    sheet["A16"].font = Font(name="Inter", size=12, bold=True, color=LODBLA)

    sheet["A17"] = "Förslag (AI)"
    sheet["B17"] = "= SUM(Anbud!F:F)"
    sheet["A18"] = "Justerat (efter granskning)"
    sheet["B18"] = "= SUM(Anbud!J:J)"

    for r in [17, 18]:
        sheet.cell(row=r, column=1).font = Font(name="Inter", size=10, color=JARN)
        c = sheet.cell(row=r, column=2)
        c.font = Font(name="JetBrains Mono", size=11, bold=True, color=LODBLA)
        c.number_format = '#,##0 "kr"'
        c.alignment = Alignment(horizontal="right")

    sheet["A21"] = "Att granska innan inlämning"
    sheet["A21"].font = Font(name="Inter", size=12, bold=True, color=LODBLA)

    todos = [
        "Kontrollera alla rader med låg konfidens — manuell prissättning kan krävas",
        "Verifiera UE-poster (kolumn med UE-kategori, ej i denna demo)",
        "Granska klumpsummerader och sätt rätt belopp där de är 1 kr (placeholder)",
        "Lägg på generellt påslag för marknadsläge (om relevant)",
        "Kontrollera att alla AMA-koder från förfrågan finns med",
    ]

    for idx, todo in enumerate(todos, start=22):
        sheet.cell(row=idx, column=1, value=f"☐  {todo}").font = Font(
            name="Inter", size=10, color=JARN
        )

    sheet.column_dimensions["A"].width = 45
    sheet.column_dimensions["B"].width = 18


def build_historik_sheet(wb: Workbook, doc_data: dict) -> None:
    sheet = wb.create_sheet("Historik per rad")
    sheet.sheet_view.showGridLines = False

    sheet["A1"] = "HISTORISKA MATCHNINGAR"
    sheet["A1"].font = Font(name="Inter", size=18, bold=True, color=LODBLA)

    sheet["A2"] = "Per anbudsrad: vilka tidigare offerter användes som referens?"
    sheet["A2"].font = Font(name="Inter", size=10, italic=True, color="6B6B6B")

    headers = [
        ("AMA-kod", 12),
        ("Beskrivning", 45),
        ("Match-rang", 11),
        ("Källprojekt", 30),
        ("Källdatum", 12),
        ("Hist. à-pris", 14),
        ("Indexjusterat", 14),
        ("Likhet", 10),
    ]

    for col_idx, (h, w) in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=col_idx, value=h)
        style_header_cell(cell)
        sheet.column_dimensions[get_column_letter(col_idx)].width = w

    sheet.row_dimensions[4].height = 26

    demo_projects = [
        ("Kv. Tegelbruket etapp 2, Sundsvall", "2024-03-15"),
        ("GC-väg Skogsbacken, Sollefteå", "2023-09-22"),
        ("Belysning Industrigatan, Härnösand", "2024-11-08"),
        ("Renovering Storgatan, Bollnäs", "2023-06-14"),
        ("Vägbelysning Rv 84, Ljusdal", "2025-01-20"),
    ]

    row = 5
    priced_lines = [l for l in doc_data["lines"] if l.get("unit_price") is not None][:8]

    for line in priced_lines:
        for match_idx in range(3):
            project, date = demo_projects[(match_idx + hash(line["description"])) % len(demo_projects)]
            base_price = line["unit_price"]
            variation = random.uniform(0.85, 1.15)
            historical = round(base_price * variation, 0)
            adjusted = round(historical * 1.04, 0)
            similarity = round(random.uniform(0.78, 0.96), 2)

            sheet.cell(row=row, column=1, value=line["ama_code"] if match_idx == 0 else "")
            sheet.cell(row=row, column=2, value=line["description"][:55] if match_idx == 0 else "")
            sheet.cell(row=row, column=3, value=match_idx + 1)
            sheet.cell(row=row, column=4, value=project)
            sheet.cell(row=row, column=5, value=date)
            sheet.cell(row=row, column=6, value=historical)
            sheet.cell(row=row, column=7, value=adjusted)
            sheet.cell(row=row, column=8, value=similarity)

            for col_idx in range(1, 9):
                cell = sheet.cell(row=row, column=col_idx)
                mono = col_idx in {1, 3, 6, 7, 8}
                style_data_cell(cell, mono=mono)

            sheet.cell(row=row, column=6).number_format = '#,##0 "kr"'
            sheet.cell(row=row, column=7).number_format = '#,##0 "kr"'
            sheet.cell(row=row, column=8).number_format = '0.00'

            if match_idx > 0:
                for col_idx in range(1, 9):
                    sheet.cell(row=row, column=col_idx).fill = PatternFill(
                        "solid", fgColor=LJUSGRA
                    )

            row += 1

        row += 1

    sheet.freeze_panes = "A5"


def build_workbook(doc_data: dict, generated_at: str, seed: int = 42) -> bytes:
    """Bygg en Excel-mall i minnet och returnera som bytes."""
    random.seed(seed)
    wb = Workbook()
    build_anbud_sheet(wb, doc_data, generated_at)
    build_summary_sheet(wb, doc_data)
    build_historik_sheet(wb, doc_data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
